"""
Museum Story Generator - Multi-LLM Support
Generates narrative stories for museum objects using various AI models
Includes Neo4j data extraction capability
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import time
import os
from typing import Optional, List, Dict, Any


try:
    from langchain_community.graphs import Neo4jGraph
    from py2neo import Graph
    NEO4J_AVAILABLE = True
except ImportError:
    NEO4J_AVAILABLE = False
    print("⚠️  Neo4j libraries not installed. Data extraction disabled.")
    print("   Install with: pip install langchain-community py2neo")

# LLM Configuration
LLM_CONFIGS = {
    'groq': {
        'env_key': 'TOKEN',
        'model': 'llama-3.3-70b-versatile',
        'library': 'groq'
    },
    'openai': {
        'env_key': 'TOKEN',
        'model': 'gpt-4o',
        'library': 'openai'
    },
    'gemini': {
        'env_key': 'TOKEN',
        'model': 'gemini-2.0-flash-exp',
        'library': 'google-generativeai'
    },
    'anthropic': {
        'env_key': 'TOKEN',
        'model': 'claude-sonnet-4-5-20250929',
        'library': 'anthropic'
    },
    'phi4': {
        'env_key': 'TOKEN',
        'model': 'Phi-4',
        'library': 'langchain',
        'base_url': 'https://models.inference.ai.azure.com'
    }
}

STORY_PROMPT = """You are a knowledgeable and passionate museum curator with expertise in archaeology and cultural heritage.

Below is structured information about an archaeological object from the MUDEC museum's collection. Your task is to narrate a vivid, emotionally engaging story of the object.

Please speak in a respectful, accessible tone — balancing storytelling with historical context.

GUIDELINES:
- Write for the ear, not the eye: use flowing, conversational language that sounds natural when spoken
- Use natural transitions between ideas ("Moving forward in time...", "Centuries later...", "Now picture...")
- Provide clear spatial descriptions ("roughly the size of a small table...", "smooth as polished stone")
- IMPORTANT: The final text must be between 170 and 270 words
- The text must be purely narrative. Do NOT include lists, bullet points, tables, or metadata

OBJECT INFORMATION (structured data from Neo4j knowledge graph):
{}

NARRATIVE STRUCTURE FOR AUDIO:
1. Opening Hook (25–35 words): Begin with an evocative scene or intriguing detail that immediately engages the listener
2. Physical Description (20–30 words): Help listeners visualize the object using size comparisons, material descriptions, and tactile qualities
3. Historical Relevance (50–70 words): Inform visitors about the historical background of the piece and the cultural context it comes from
4. Historical Journey (60–80 words): Tell the story chronologically with clear time markers and smooth transitions between eras
5. Present-Day Connection (25–30 words): Conclude by linking the object's journey to today, emphasizing its relevance to visitors
"""


# ========== NEO4J DATA EXTRACTION ==========

class Neo4jExtractor:
    """Extract object information from Neo4j database"""
    
    def __init__(self, url="bolt://localhost:7687", username="neo4j", password="neo4j"):
        if not NEO4J_AVAILABLE:
            raise ImportError("Neo4j libraries not installed")
        
        self.graph_langchain = Neo4jGraph(url=url, username=username, password=password)
        self.graph = Graph(url, auth=(username, password))
    
    def get_actor_name(self, actor_identifier: str) -> str:
        """Get actor name using identifier"""
        if not actor_identifier:
            return "Unknown"
        
        query = """
        MATCH (actor)
        WHERE (actor:E39_Actor OR actor:E21_Person OR actor:E40_Legal_Body)
          AND (actor.id = $identifier OR actor.name = $identifier)
        OPTIONAL MATCH (actor)-[:P1_is_identified_by]->(app:E41_Appellation)
        RETURN app.value
        LIMIT 1
        """
        
        result = self.graph.run(query, identifier=actor_identifier).data()
        
        if result and result[0] and result[0].get('app.value'):
            return result[0]['app.value']
        
        return actor_identifier
    
    def get_actor_names_batch(self, actor_ids: List[str], actor_names: List[str]) -> Dict[str, str]:
        """Get actor names in batch"""
        name_map = {}
        
        for actor_id, actor_name in zip(actor_ids, actor_names):
            identifier = actor_id or actor_name
            if identifier and identifier not in name_map:
                name_map[identifier] = self.get_actor_name(identifier)
        
        return name_map
    
    def get_object_info(self, local_id: str) -> Optional[Dict[str, Any]]:
        """Retrieve all information for an object"""
        query = """
        MATCH (obj:E22_Man_Made_Object {localId: $local_id})
        
        OPTIONAL MATCH (obj)-[:P45_consists_of]->(material:E57_Material)
        OPTIONAL MATCH (obj)-[:P32_used_general_technique]->(tech:E29_Design_or_Procedure)
        OPTIONAL MATCH (obj)-[:P43_has_dimension]->(dim:E54_Dimension)
        OPTIONAL MATCH (obj)-[:P2_has_type]->(typ:E55_Type)
        OPTIONAL MATCH (obj)-[:P52_has_current_owner]->(owner:E39_Actor)
        OPTIONAL MATCH (obj)-[:P1_is_identified_by]->(identifier:E42_Identifier)
        OPTIONAL MATCH (obj)-[:P1_is_identified_by]->(appellation:E41_Appellation)
        OPTIONAL MATCH (period:E4_Period)-[:P12_occurred_in_the_presence_of]->(obj)
        
        OPTIONAL MATCH (obj)-[:P108i_was_produced_by]->(prod:E12_Production)
        OPTIONAL MATCH (prod)-[:P4_has_time_span]->(prodTime:E52_Time_Span)
        OPTIONAL MATCH (prod)-[:P7_took_place_at]->(prodPlace:E53_Place)

        RETURN obj,
               collect(DISTINCT material.label) AS materials,
               collect(DISTINCT tech.label) AS techniques,
               collect(DISTINCT {type: dim.type, value: dim.value}) AS dimensions,
               collect(DISTINCT typ.label) AS types,
               collect(DISTINCT owner.name) AS current_owners,
               collect(DISTINCT identifier.value) AS identifiers,
               collect(DISTINCT {label: appellation.label, value: appellation.value}) AS appellations,
               collect(DISTINCT period.label) AS periods,
               collect(DISTINCT {
                   id: prod.id,
                   type: prod.type,
                   time: {
                       value: prodTime.value,
                       begin: prodTime.P82a_begin_of_the_begin,
                       end: prodTime.P82b_end_of_the_end
                   },
                   place: prodPlace.name
               }) AS production_data
        """
        
        result = self.graph.run(query, local_id=local_id).data()
        if not result or not result[0].get('obj'):
            return None
        
        obj_info = result[0]
        
        # Acquisition events
        acq_query = """
        MATCH (obj:E22_Man_Made_Object {localId: $local_id})
        MATCH (acq:E8_Acquisition)-[:P24_transferred_title_of]->(obj)
        
        OPTIONAL MATCH (acq)-[:P4_has_time_span]->(acqTime:E52_Time_Span)
        OPTIONAL MATCH (acq)-[:P22_transferred_title_to]->(acqTo)
        OPTIONAL MATCH (acq)-[:P23_transferred_title_from]->(acqFrom)
        OPTIONAL MATCH (acq)-[:P14_carried_out_by]->(acqBy)
        
        RETURN acq.id as id,
               acq.type as type,
               {
                   value: acqTime.value,
                   begin: acqTime.P82a_begin_of_the_begin,
                   end: acqTime.P82b_end_of_the_end
               } as time,
               acqTo.id as to_id,
               acqTo.name as to_name,
               acqFrom.id as from_id,
               acqFrom.name as from_name,
               acqBy.id as by_id,
               acqBy.name as by_name
        ORDER BY acq.id
        """
        
        acq_results = self.graph.run(acq_query, local_id=local_id).data()
        
        # Batch fetch actor names
        actor_ids = []
        actor_names = []
        for row in acq_results:
            actor_ids.extend([row.get('to_id'), row.get('from_id'), row.get('by_id')])
            actor_names.extend([row.get('to_name'), row.get('from_name'), row.get('by_name')])
        
        actor_name_map = self.get_actor_names_batch(actor_ids, actor_names)
        
        # Process acquisitions
        acquisition_data = []
        for row in acq_results:
            to_key = row.get('to_id') or row.get('to_name')
            from_key = row.get('from_id') or row.get('from_name')
            by_key = row.get('by_id') or row.get('by_name')
            
            acquisition_data.append({
                'id': row.get('id'),
                'type': row.get('type'),
                'time': row.get('time'),
                'to_name': actor_name_map.get(to_key, 'Unknown') if to_key else 'Unknown',
                'from_name': actor_name_map.get(from_key, 'Unknown') if from_key else 'Unknown',
                'by_name': actor_name_map.get(by_key, 'Unknown') if by_key else 'Unknown'
            })
        
        obj_info['acquisition_data'] = acquisition_data
        obj_info['production_data'] = [p for p in obj_info.get('production_data', []) if p.get('id')]
        
        return obj_info
    
    def describe_object(self, obj_info: Dict[str, Any]) -> str:
        """Convert object info to formatted description"""
        if not obj_info or "obj" not in obj_info:
            return "No information available for this object."

        obj = obj_info["obj"]
        local_id = obj.get("localId", "Unknown")

        def clean_list(items):
            return [str(i) for i in items if i] if items else []

        identifiers = clean_list(obj_info.get("identifiers", []))
        
        appellations = []
        for app in obj_info.get("appellations", []):
            if isinstance(app, dict):
                appellations.append(app.get("label") or app.get("value"))
            elif app:
                appellations.append(str(app))
        appellations = [a for a in appellations if a]
        
        materials = clean_list(obj_info.get("materials", []))
        techniques = clean_list(obj_info.get("techniques", []))
        types = clean_list(obj_info.get("types", []))
        owners = clean_list(obj_info.get("current_owners", []))
        periods = clean_list(obj_info.get("periods", []))
        
        dimensions = []
        for d in obj_info.get("dimensions", []):
            if isinstance(d, dict) and d.get("type") and d.get("value"):
                dimensions.append(f'{d["type"]}: {d["value"]}')

        productions = []
        for prod in obj_info.get("production_data", []):
            if not prod or not prod.get('id'):
                continue
            
            time_data = prod.get('time', {})
            if isinstance(time_data, dict):
                time_str = time_data.get('value') or f"{time_data.get('begin', '?')} - {time_data.get('end', '?')}"
            else:
                time_str = "Unknown"
            
            place_str = prod.get('place') or "Unknown"
            
            productions.append(
                f"  • ID: {prod.get('id')}, Type: {prod.get('type')}\n"
                f"    Time: {time_str}\n"
                f"    Place: {place_str}"
            )

        acquisitions = []
        for acq in obj_info.get("acquisition_data", []):
            if not acq or not acq.get('id'):
                continue
            
            time_data = acq.get('time', {})
            if isinstance(time_data, dict):
                time_str = time_data.get('value') or f"{time_data.get('begin', '?')} - {time_data.get('end', '?')}"
            else:
                time_str = "Unknown"
            
            event_type = acq.get('type', '')
            
            event_lines = [
                f"  • ID: {acq.get('id')}, Type: {event_type}",
                f"    Time: {time_str}"
            ]
            
            if acq.get('from_name') and acq.get('from_name') != 'Unknown':
                event_lines.append(f"    Transferred from: {acq.get('from_name')}")
            
            if acq.get('to_name') and acq.get('to_name') != 'Unknown':
                event_lines.append(f"    Transferred to: {acq.get('to_name')}")
            
            if event_type in ['Creation', 'Collection'] and acq.get('by_name') and acq.get('by_name') != 'Unknown':
                event_lines.append(f"    Carried out by: {acq.get('by_name')}")
            
            acquisitions.append('\n'.join(event_lines))

        return f"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Object: {local_id}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📋 IDENTIFIERS
{', '.join(identifiers) if identifiers else '  Not recorded'}

🏷️ NAMES/TITLES
{', '.join(appellations) if appellations else '  Not recorded'}

🎨 CLASSIFICATION
  Type: {', '.join(types) if types else 'Not recorded'}

🧱 PHYSICAL PROPERTIES
  Materials: {', '.join(materials) if materials else 'Not recorded'}
  Techniques: {', '.join(techniques) if techniques else 'Not recorded'}
  Dimensions: {'; '.join(dimensions) if dimensions else 'Not recorded'}

👤 OWNERSHIP
  Current Owner: {', '.join(owners) if owners else 'Not recorded'}

🕰️ HISTORICAL CONTEXT
  Periods: {', '.join(periods) if periods else 'Not recorded'}

🏭 PRODUCTION EVENTS
{chr(10).join(productions) if productions else '  None recorded'}

📦 ACQUISITION EVENTS
{chr(10).join(acquisitions) if acquisitions else '  None recorded'}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
""".strip()
    
    def get_all_object_ids(self, limit: int = 100) -> List[str]:
        """Get all object IDs from database"""
        query = f"""
        MATCH (obj:E22_Man_Made_Object)
        RETURN obj.localId as ID
        LIMIT {limit}
        """
        
        result = self.graph_langchain.query(query)
        return [item['ID'] for item in result if item['ID']]
    
    def extract_to_excel(self, output_file: str = "catalog_100_objects.xlsx", limit: int = 100):
        """Extract object data to Excel file"""
        print("🔍 Retrieving object IDs...")
        object_ids = self.get_all_object_ids(limit=limit)
        print(f"✅ Found {len(object_ids)} objects\n")
        
        data = []
        
        for i, local_id in enumerate(object_ids, 1):
            print(f"📦 Processing object {i}/{len(object_ids)}: {local_id}")
            
            try:
                obj_info = self.get_object_info(local_id)
                
                if obj_info:
                    description = self.describe_object(obj_info)
                    data.append({
                        'N°': i,
                        'Object ID': local_id,
                        'Description': description
                    })
                else:
                    data.append({
                        'N°': i,
                        'Object ID': local_id,
                        'Description': 'No information available'
                    })
                    
            except Exception as e:
                print(f"  ⚠️  Error for object {local_id}: {e}")
                data.append({
                    'N°': i,
                    'Object ID': local_id,
                    'Description': f'Error retrieving information: {str(e)}'
                })
        
        df = pd.DataFrame(data)
        
        print(f"\n💾 Saving to {output_file}...")
        df.to_excel(output_file, index=False, engine='openpyxl')
        
        print("🎨 Formatting...")
        wb = load_workbook(output_file)
        ws = wb.active
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 100
        
        for cell in ws[1]:
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=3)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='top')
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='top')
            
            ws.row_dimensions[row].height = 300
        
        wb.save(output_file)
        
        print(f"\n✅ Excel file created: {output_file}")
        print(f"📊 Objects: {len(data)}")
        
        return output_file


class StoryGenerator:
    def __init__(self, provider='groq', model=None, api_key=None):
        self.provider = provider.lower()
        self.config = LLM_CONFIGS.get(self.provider)
        
        if not self.config:
            raise ValueError(f"Unknown provider: {provider}. Choose from: {list(LLM_CONFIGS.keys())}")
        
        self.api_key = api_key or os.getenv(self.config['env_key'])
        if not self.api_key:
            raise ValueError(f"API key not found. Set TOKEN env variable or pass api_key parameter")
        
        self.model = model or self.config['model']
        self.client = self._init_client()
    
    def _init_client(self):
        if self.config['library'] == 'groq':
            from groq import Groq
            return Groq(api_key=self.api_key)
        
        elif self.config['library'] == 'openai':
            from openai import OpenAI
            return OpenAI(api_key=self.api_key)
        
        elif self.config['library'] == 'langchain':
            from langchain_openai import ChatOpenAI
            return ChatOpenAI(
                model=self.model,
                api_key=self.api_key,
                base_url=self.config.get('base_url'),
                temperature=0.7
            )
        
        elif self.config['library'] == 'google-generativeai':
            import google.generativeai as genai
            genai.configure(api_key=self.api_key)
            return genai.GenerativeModel(self.model)
        
        elif self.config['library'] == 'anthropic':
            from anthropic import Anthropic
            return Anthropic(api_key=self.api_key)
        
        else:
            raise ValueError(f"Unknown library: {self.config['library']}")
    
    def generate(self, description, retries=3):
        if not description or description.strip() == '' or description == 'No information available':
            return "Insufficient information to generate a narrative for this object."
        
        prompt = STORY_PROMPT.format(description)
        
        for attempt in range(retries):
            try:
                if self.config['library'] == 'groq':
                    response = self.client.chat.completions.create(
                        model=self.model,
                        messages=[{"role": "user", "content": prompt}],
                        max_tokens=500,
                        temperature=0.7
                    )
                    return response.choices[0].message.content.strip()
                
                elif self.config['library'] == 'openai':
                    response = self.client.chat.completions.create(
                        model=self.model,
                        messages=[{"role": "user", "content": prompt}],
                        max_tokens=500,
                        temperature=0.7
                    )
                    return response.choices[0].message.content.strip()
                
                elif self.config['library'] == 'langchain':
                    from langchain.schema import HumanMessage, SystemMessage
                    messages = [
                        SystemMessage(content="You are an expert museum curator specializing in archaeological storytelling."),
                        HumanMessage(content=prompt)
                    ]
                    response = self.client.invoke(messages)
                    return response.content.strip()
                
                elif self.config['library'] == 'google-generativeai':
                    response = self.client.generate_content(prompt)
                    return response.text.strip()
                
                elif self.config['library'] == 'anthropic':
                    message = self.client.messages.create(
                        model=self.model,
                        max_tokens=1024,
                        temperature=0.7,
                        system="You are an expert museum curator specializing in archaeological storytelling.",
                        messages=[{"role": "user", "content": prompt}]
                    )
                    return message.content[0].text.strip()
                
            except Exception as e:
                error = str(e)
                print(f"    ⚠️  Attempt {attempt + 1}/{retries} failed: {error[:100]}")
                
                if "rate_limit" in error.lower() or "429" in error:
                    print(f"    ⏳ Rate limited, waiting 60s...")
                    time.sleep(60)
                elif "timeout" in error.lower():
                    time.sleep(10)
                elif attempt < retries - 1:
                    time.sleep(5)
                else:
                    return f"Error after {retries} attempts: {error[:200]}"
        
        return "Failed to generate story after multiple attempts."


def process_catalog(input_file, output_file, provider='groq', model=None, 
                   api_key=None, save_interval=10):
    
    print(f"📖 Reading {input_file}...")
    print(f"🤖 Provider: {provider}")
    if model:
        print(f"🎯 Model: {model}")
    print()
    
    try:
        df = pd.read_excel(input_file)
    except Exception as e:
        print(f"❌ Error reading file: {e}")
        return
    
    print(f"✅ Found {len(df)} objects\n")
    
    generator = StoryGenerator(provider=provider, model=model, api_key=api_key)
    
    if 'Narrative Story' not in df.columns:
        df['Narrative Story'] = ''
    
    existing = df['Narrative Story'].notna().sum()
    if existing > 0:
        print(f"📊 {existing} stories already exist\n")
    
    total = len(df)
    success = 0
    errors = 0
    start = time.time()
    
    for idx, row in df.iterrows():
        if pd.notna(row['Narrative Story']) and row['Narrative Story'] != '':
            success += 1
            print(f"⏭️  {idx + 1}/{total}: {row['Object ID']} - Skip")
            continue
        
        obj_id = row['Object ID']
        desc = row['Description']
        
        print(f"✍️  {idx + 1}/{total}: {obj_id}")
        
        try:
            story = generator.generate(desc)
            df.at[idx, 'Narrative Story'] = story
            
            if "Error" in story or "Failed" in story or "Insufficient" in story:
                errors += 1
                print(f"    ⚠️  Generation issue")
            else:
                success += 1
                wc = len(story.split())
                print(f"    ✅ Success ({wc} words)")
        
        except Exception as e:
            df.at[idx, 'Narrative Story'] = f"Error: {str(e)}"
            errors += 1
            print(f"    ❌ {e}")
        
        if (idx + 1) % save_interval == 0:
            elapsed = time.time() - start
            avg = elapsed / (idx + 1 - existing) if (idx + 1 - existing) > 0 else 0
            remaining = (total - idx - 1) * avg if avg > 0 else 0
            
            print(f"  💾 Checkpoint ({idx + 1}/{total})")
            print(f"  ⏱️  {avg:.1f}s/obj - {remaining/60:.1f}m left")
            
            try:
                df.to_excel(output_file, index=False)
                print(f"  ✅ Saved - Success: {success}, Errors: {errors}\n")
            except Exception as e:
                print(f"  ⚠️  Save error: {e}\n")
        
        time.sleep(1)
    
    print(f"\n💾 Final save to {output_file}...")
    try:
        df.to_excel(output_file, index=False, engine='openpyxl')
        
        wb = load_workbook(output_file)
        ws = wb.active
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 100
        ws.column_dimensions['D'].width = 80
        
        for cell in ws[1]:
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=3).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws.cell(row=r, column=4).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='top')
            ws.cell(row=r, column=2).alignment = Alignment(horizontal='left', vertical='top')
            ws.row_dimensions[r].height = 300
        
        wb.save(output_file)
        
        total_time = time.time() - start
        
        print(f"\n✅ Complete!")
        print(f"📊 Summary:")
        print(f"   Total: {total}")
        print(f"   Success: {success}")
        print(f"   Errors: {errors}")
        print(f"   Success rate: {(success/total)*100:.1f}%")
        print(f"   Time: {total_time/60:.1f}m")
        if (total - existing) > 0:
            print(f"   Avg: {total_time/(total-existing):.1f}s/obj")
        
    except Exception as e:
        print(f"❌ Save error: {e}")


if __name__ == "__main__":
   
    # Uncomment and configure to use:
    """
    extractor = Neo4jExtractor(
        url="bolt://localhost:7687",
        username="neo4j",
        password="your_password"
    )
    
    # Extract object data to Excel
    catalog_file = extractor.extract_to_excel(
        output_file="catalog_100_objects.xlsx",
        limit=100
    )
    
    # Generate stories for the extracted data
    process_catalog(
        input_file=catalog_file,
        output_file="catalog_with_stories.xlsx",
        provider='groq',
        save_interval=10
    )
    """
    
   
    process_catalog(
        input_file="catalog_100_objects.xlsx",
        output_file="catalog_with_stories.xlsx",
        provider='groq',  
        model=None,  
        save_interval=10
    )
